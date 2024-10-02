import pdfplumber
import pandas as pd
from openpyxl import load_workbook

# Caminhos dos arquivos
pdf_path = "c:\\Users\\jhennifer.nascimento\\nfs\\pdf\\relatorio.pdf"  # Caminho para o PDF
excel_path = "c:\\Users\\jhennifer.nascimento\\nfs\\relatorio.xlsx"  # Caminho para o Excel

# Abre o PDF e carrega o Excel
pdf = pdfplumber.open(pdf_path)
wb = load_workbook(excel_path)
ws = wb.active


# Adiciona cabeçalhos na planilha Excel se não houver
header = ["Código", "Descrição", "Grupo de Produto", "Unidade", "Tipo do Produto", "NCM", "Cod. Barras", "Estoque Mínimo", "Ponto de Pedido", "Estoque Máximo"]
if ws.max_row == 1 and ws.cell(row=1, column=1).value is None:
    ws.append(header)

# Função para extrair os dados de um bloco de texto
def extrair_dados_bloco(bloco):
    dados = {}
    
    # Lógica para extrair as informações de cada linha
    linhas = bloco.split("\n")
    
    for linha in linhas:
        if "Código:" in linha and "NCM:" in linha:
            partes = linha.split("NCM:")
            dados["Código"] = partes[0].replace("Código:", "").strip()
            dados["NCM"] = partes[1].strip()
        elif "Descrição:" in linha and "Cod. Barras:" in linha:
            partes = linha.split("Cod. Barras:")
            dados["Descrição"] = partes[0].replace("Descrição:", "").strip()
            dados["Cod. Barras"] = partes[1].strip()
        elif "Grupo de Produto:" in linha and "Estoque Mínimo:" in linha:
            partes = linha.split("Estoque Mínimo:")
            dados["Grupo de Produto"] = partes[0].replace("Grupo de Produto:", "").strip()
            dados["Estoque Mínimo"] = partes[1].strip()
        elif "Unidade:" in linha and "Estoque Máximo:" in linha:
            partes = linha.split("Estoque Máximo:")
            dados["Unidade"] = partes[0].replace("Unidade:", "").strip()
            dados["Estoque Máximo"] = partes[1].strip()
        elif "Tipo do Produto:" in linha and "Ponto de Pedido:" in linha:
            partes = linha.split("Ponto de Pedido:")
            dados["Tipo do Produto"] = partes[0].replace("Tipo do Produto:", "").strip()
            dados["Ponto de Pedido"] = partes[1].strip()

    return dados

# Função para extrair todos os blocos de dados de uma página
def extrair_dados_pagina(texto):
    produtos = []
    # Divide o texto usando a linha pontilhada como separador
    blocos = texto.split("                     ")
    for bloco in blocos:
        bloco = bloco.strip()  # Remove espaços em branco extras
        if bloco:  # Verifica se o bloco não está vazio
            dados = extrair_dados_bloco(bloco)
            if dados:
                produtos.append(dados)
    return produtos

# Percorre todas as páginas do PDF
for page_number, page in enumerate(pdf.pages, start=1):
    texto = page.extract_text()
    
    # Debug: Exibe o texto extraído da página
    print(f"Página {page_number}:\n{texto}\n{'-'*40}")
    
    # Extrai os dados de todos os produtos na página
    produtos = extrair_dados_pagina(texto)
    
    # Debug: Exibe os produtos extraídos
    print(f"Produtos extraídos da página {page_number}: {produtos}\n{'='*40}")
    
    # Adiciona os dados à planilha
    for dados in produtos:
        ws.append([
            dados.get("Código", ""),
            dados.get("Descrição", ""),
            dados.get("Grupo de Produto", ""),
            dados.get("Unidade", ""),
            dados.get("Tipo do Produto", ""),
            dados.get("NCM", ""),
            dados.get("Cod. Barras", ""),
            dados.get("Estoque Mínimo", ""),
            dados.get("Ponto de Pedido", ""),
            dados.get("Estoque Máximo", "")
        ])

# Salva a planilha Excel preenchida
wb.save("c:\\Users\\jhennifer.nascimento\\nfs\\result.xlsx")
pdf.close()
